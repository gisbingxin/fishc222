from scipy import io
import matplotlib.pyplot as plt
import numpy as np
import xlrd
import openpyxl
from osgeo import gdal
import matlab
import matlab.engine
import self_strech_img
import self_read_position_excel

def read_show_img(image_name, show_img=False):
    # raster = gdal.Open('C:\\test.jpg')
    # image_name = 'C:\hyperspectral\AVIRISReflectanceSubset.dat'
    raster = gdal.Open(image_name)
    # raster = gdal.Open('e:\qwer.img')
    print('Reading the image named:', image_name)
    # print('geotrans is ', raster.GetGeoTransform())
    # print('projection is ',raster.GetProjection())
    xsize = raster.RasterXSize  # RasterXSize是图像宽
    ysize = raster.RasterYSize  # RasterYSize是图像的高
    band_num = raster.RasterCount  # RasterCount是图像的波段数
    im_GeoTrans = raster.GetGeoTransform() #仿射矩阵，左上角像素的大地坐标和像素分辨率
                        # (左上角x, x分辨率，仿射变换，左上角y坐标，y分辨率，仿射变换)
    im_proj = raster.GetProjection() #地图投影信息，字符串表示
    print('in read and show x,y', xsize, ysize)

    # 获取rgb对应波段数值，传递给拉伸显示模块

    raster_array = raster.ReadAsArray()  # raster_array.shape=(191,128,155)对应（band_num,Ysize,Xsize),
    # Ysize代表行数，Xsize代表列数
    raster_array_np = np.array(raster_array)

    #
    # print('rnp',np.shape(rnp))
    print('raster_array shape in read_show img', raster_array_np.shape)
    if band_num >= 3:
        r = raster_array[0]
        g = raster_array[1]
        b = raster_array[2]
        print('type of r in read_show_img:', type(r))

        if show_img:  # 判断是否需要显示图像
            strechimg(xsize, ysize, r, g, b)  # 调用图像拉伸显示模块
            # strechimg(92, 92, r[0:92,0:92], g[0:92,0:92], b[0:92,0:92])  # 调用图像拉伸显示模块
    else:
        img_data = raster.GetRasterBand(1)
        img_data = img_data.ReadAsArray()
        # print(np.shape(img_data))
        if show_img:
            self_strech_img.strechimg_pan(xsize, ysize, img_data)
            # print()
    return raster, raster_array, xsize, ysize, band_num, im_GeoTrans, im_proj
def get_1D_sample_data(raster, band_num, class_num, num_per_class, total_num,sample_num,
                       excel_name, sheet_num, start_row, start_col, end_row, end_col,
                       sample_size_X=1, sample_size_Y=1, channel_1D=1, random=False):
    # 读取Excel中存储的采样点位置数据，返回值中sample_num代表类别编号
    if random:
        sample_pos = self_read_position_excel.read_sample_position_random\
            (excel_name, sheet_num, num_per_class, total_per_class, start_col,
                                                 start_row)
    else:
        sample_pos = self_read_position_excel.read_sample_position\
            (excel_name, sheet_num, start_row, start_col, end_row, end_col)

        # 该函数返回所选择的sample的位置点，是一个三维数组shape为[class_num,X_num,Y_num]

    # #sample定义为[batch，band,Ysize,Xsize],batch这里是所有类别采样点的数量
    # x_data = np.zeros([sample_num, band_num, sample_size_Y, sample_size_X], dtype=float)
    x_data = np.zeros([sample_num, band_num, channel_1D], dtype=float)
    # x_data = []
    # 用来存储各类别各采样点的图像灰度值
    y_data = np.zeros([sample_num, class_num], dtype=int)

    margin_pixel = []
    loc_origin = 0
    for i in range(0, class_num):  # i表示类的序号
        for x_i in range(loc_origin, loc_origin + num_per_class[i]):  # x_i 表示在该类内的序号，即行号
            # for y_i in range(0,2):
            x_offset = int(sample_pos[x_i, 0]) - 1  # GDAL中的ReadAsArray传入参数应当为int型
            # 而此处sample是numpy.int32，所以需要进行数据类型转换
            y_offset = int(sample_pos[x_i, 1]) - 1
            # print('x,y location:', x_locate,y_locate)
            data_from_raster = raster.ReadAsArray(x_offset, y_offset, sample_size_X, sample_size_Y)
            # print('shape of data_from_raster',np.shape(data_from_raster))
            data = np.swapaxes(data_from_raster, 0, 1)
            x_data[x_i, :, :] = data
            # y_data[num_per_class[i]*i+x_i,i] = 1             #与x_data对应的batch处，赋值为1，其余位置为0
            y_data[x_i, i] = 1
        loc_origin = loc_origin + num_per_class[i]
        # print('shape of x_data in get next batch',np.shape(x_data),np.shape(y_data))
    return x_data, y_data

def excel_save(xdata,ydata,class_num,band_num,filename):
    # Try to save a list variable in txt file.
    wb = openpyxl.Workbook()
    #ew = openpyxl.ExcelWriter(workbook=wb)
    sheet_new= wb.create_sheet('new sheet')
    # sheet_new1 = wb.add_sheet('new sheet1',cell_overwrite_ok=True)
    # sheet_new2 = wb.add_sheet('new sheet2', cell_overwrite_ok=True)
    # sheet_new3 = wb.add_sheet('new sheet3', cell_overwrite_ok=True)
    #start=0
    for row_i in range(0,class_num*150):
        yy=ydata[row_i]
        sheet_new.cell(row_i+1,1).value=yy

    for row_i in range(0,class_num*150):#height*width):--->y
        #row_num= int(row_i/height)
        for col_i in range(0,band_num):#---->x
            cell_id=row_i*class_num+col_i
            xx = float(xdata[row_i, col_i])
            # sheet_new.cell(cell_id+1,1).value=row_i+1
            # sheet_new.cell(cell_id + 1, 2).value = col_i+1
            sheet_new.cell(row_i+1, col_i+2).value = xx
        #print(row_i)
    wb.save(filename)


if __name__=='__main__':
    image_name = 'G:\data for manuscripts\AVIRIS20100517\\' \
                 'f100517t01p00r10 For CNN\\f100517t01p00r10rdn_b_sc01_ort_img_resized2radiance_resized2_flaashed_deletebands_eq0.img'
    excel_name = 'G:\data for manuscripts\AVIRIS20100517\\fig_thickness_resized_roated_class_resize_ROIs\ROIs.xlsx'
    num_per_class = np.array([150, 150, 150, 150, 150, 150, 150])
    sample_num = np.sum(num_per_class)  # class_num * num_per_class  #训练数据中，所有类采样点的总数。对应后面的batch
    total_per_class=np.array([196, 1485, 195, 293, 196, 199, 236])
    sample_size_X=1
    sample_size_Y=1


    start_row = 1  # 表示记录采样点数据的Excel中，数据开始的行，0表示第一行
    end_row = start_row + num_per_class - 1

    start_col = 1  # 表示记录采样点数据的Excel中，数据开始的列，0表示第一列
    end_col = 2  # 如果行列数字错误，可能出现如下错误：
    # ERROR 5: Access window out of range in RasterIO().  Requested
    # (630,100) of size 10x10 on raster of 634x478.
    class_num=sheet_num = 7  # 表示Excel中sheet的数目，必须与类别数量一致

    show_img = False  # 用于判断是否对图像进行显示
    raster, raster_array, xsize, ysize, band_num, im_geotrans, im_proj = read_show_img(image_name, show_img)
    x_data, y_data = get_1D_sample_data(raster, band_num, sheet_num, num_per_class, total_per_class,sample_num,
                                        excel_name, sheet_num, start_row, start_col, end_row, end_col,
                                        sample_size_X, sample_size_Y)
    x_data=np.squeeze(x_data)
    y_data=np.where(y_data==1)
    #print('y_data',y_data1[1])

    save_path='G:\data for manuscripts\AVIRIS20100517\\f100517t01p00r10 For CNN\\xdata.xlsx'
    excel_save(x_data,y_data[1],class_num,band_num,save_path)



