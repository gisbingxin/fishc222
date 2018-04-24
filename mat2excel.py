from scipy import io
import matplotlib.pyplot as plt
import numpy as np
import xlrd
import openpyxl

def readMat(mat_path,key_w,only_get_key):
    mat = io.loadmat(mat_path) #返回值是一个字典
    if only_get_key ==True:
        print(mat.items())
        return 0,0,0,0
    im_data = mat.get(key_w)

    im_data = np.array(im_data, np.float32)
    # print(np.max(im_data))
    # print(im_data.shape)
    if len(im_data.shape) ==3:
        height,width,band_num = im_data.shape   # 注意，band_num, height,width的位置，要根据具体数据存储顺序进行改变。
                                                # 如果顺序错误，则在下面显示部分出现异常显示，不报错
    else:
        band_num,(height,width) = 1,im_data.shape

    return im_data,band_num,height,width

def excel_save(im_data,height,width,filename):
    # Try to save a list variable in txt file.
    wb = openpyxl.Workbook()
    #ew = openpyxl.ExcelWriter(workbook=wb)
    sheet_new= wb.create_sheet('new sheet')
    # sheet_new1 = wb.add_sheet('new sheet1',cell_overwrite_ok=True)
    # sheet_new2 = wb.add_sheet('new sheet2', cell_overwrite_ok=True)
    # sheet_new3 = wb.add_sheet('new sheet3', cell_overwrite_ok=True)
    for row_i in range(0,height):#height*width):
        #row_num= int(row_i/height)
        for col_i in range(0,width):
            cell_id=row_i*height+col_i
            xx = int(im_data[row_i, col_i])
            sheet_new.cell(cell_id+1,1).value=row_i+1
            sheet_new.cell(cell_id + 1, 2).value = col_i+1
            sheet_new.cell(cell_id + 1, 3).value = xx
        #print(row_i)
    wb.save(filename)

if __name__ == '__main__':
    mat_path = 'F:\Python\workshop\data\hydata\PaviaU_gt.mat'
    key_w = 'paviaU_gt'
    only_get_key =False #通过这个“开关”控制是否仅仅查看mat文件中的key值。
    im_data,band_num,height,width=readMat(mat_path,key_w,only_get_key)  #读取Mat文件并返回图像参数，
                                # 注意，band_num, height,width的位置，要根据具体数据存储顺序进行改变。
                                # 如果顺序错误，则在下面显示部分出现异常显示，不报错
    if only_get_key == False:
        filename = 'F:\\test1.xlsx'
        excel_save(im_data, height, width, filename)
